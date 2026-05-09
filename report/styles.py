# -*- coding: utf-8 -*-
"""
styles.py - Excel 样式定义
============================
颜色、字体、边框、填充等样式工具函数。
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# 样式
# ============================================================

def _mkstyle():
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return dict(
        border=border,
        hdr_fill=PatternFill("solid", fgColor="1976D2"),
        hdr_font=Font(bold=True, color="FFFFFF", size=10),
        title_font=Font(bold=True, size=13, color="212121"),
        sub_font=Font(size=9, color="757575"),
        pass_fill=PatternFill("solid", fgColor="C8E6C9"),
        fail_fill=PatternFill("solid", fgColor="FFCDD2"),
        skip_fill=PatternFill("solid", fgColor="F5F5F5"),
        alt_fill=PatternFill("solid", fgColor="FAFAFA"),
    )


def _rfill(result: str, s: dict):
    return {"PASS": s["pass_fill"], "FAIL": s["fail_fill"], "SKIP": s["skip_fill"]}.get(result.upper())


def _rfont(result: str):
    """返回测试结论单元格字体颜色：FAIL红/PASS绿/SKIP黄/NA灰，无背景填充也显示颜色。"""
    return {
        "FAIL": Font(color="CC0000", bold=True),
        "PASS": Font(color="008800", bold=True),
        "SKIP": Font(color="FF8F00", bold=True),
        "NA":   Font(color="757575"),
    }.get(result.upper(), Font(color="757575"))


def _fmt(v):
    if v == "" or v is None:
        return ""
    try:
        return f"{float(v):.3f}"
    except (ValueError, TypeError):
        return str(v)
