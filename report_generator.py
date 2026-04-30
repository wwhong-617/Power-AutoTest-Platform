#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
报告生成器 - Excel 表格报告
==========================
结构：
  - Sheet "测试结果汇总"：总体通过/失败/分类统计
  - Sheet "输入测试"/...：每分类一个 Sheet
      - 数据行（先写完所有）
      - 波形图片（后统一放置，4张/行）
依赖：openpyxl, Pillow
"""

import os
import json as _json
import uuid
import logging
from datetime import datetime
from collections import defaultdict

logger = logging.getLogger("PowerAutoTest.ReportGen")

# 从 test_engine 导入用例注册表（唯一数据源）
from test_engine import TestEngine

# ============================================================
# 映射表（统一由 TestEngine.CASE_CN_NAMES 派生）
# ============================================================

# 英文 key -> 中文显示名（只取 tuple 第0项，兼容旧版字符串）
def _display_name(en: str) -> str:
    v = TestEngine.CASE_CN_NAMES.get(en, en)
    return v[0] if isinstance(v, tuple) else v

# 英文 key -> 中文名（短名，用于内部 key）
CASE_NAME_CN_MAP = {en: _display_name(en) for en in TestEngine.CASE_CN_NAMES}

# 中文显示名（含"测试"后缀）-> 分类
_CASE_CN_TO_CATEGORY = {
    "输入":    "输入测试",
    "输出":    "输出测试",
    "过流":    "保护功能测试",
    "过压":    "保护功能测试",
    "欠压":    "保护功能测试",
    "过温":    "保护功能测试",
    "短路":    "保护功能测试",
    "协议":    "协议测试",
}

def _cat(cn: str) -> str:
    """根据中文名推断分类"""
    for prefix, category in _CASE_CN_TO_CATEGORY.items():
        if cn.startswith(prefix):
            return category
    return "其他"

CASE_TO_CATEGORY = {en: _cat(cn) for en, cn in CASE_NAME_CN_MAP.items()}


def _cn(en: str) -> str:
    return CASE_NAME_CN_MAP.get(en, en)


# 列头名称即 sub_result 的字段 key，_flatten 时直接透传
# ============================================================

DEFAULT_COLS = [
    ("序号",      6),
    ("输入电压",  12),
    ("频率",       8),
    ("协议",      12),
    ("输出电压",  12),
    ("输出电流",  12),
    ("规格上限",  13),
    ("规格下限",  13),
    ("最大值",    13),
    ("有效值",    13),
    ("最小值",    13),
    ("测试结论",  13),
    ("测试波形",  12),
    ("备注",      32),
]

# ============================================================
# 统一报告列定义
#
# 字段名（列头）= sub_result 的 key 名，两者必须完全一致
# 统一列定义，所有用例共享，用例不用的字段留空
#
# 格式：[("列头", 宽度), ...]
# ============================================================
#
# 格式：[("列头", 宽度), ...]
# GLOBAL_COLS = 汇总 Sheet（分类统计）用的表头
# 列顺序不影响各 case 详情 Sheet（详情 Sheet 用各自 COLS 定义）
# ============================================================
GLOBAL_COLS = [
    # === 标识（2列）===
    ("序号",               6),
    ("用例名称",          18),

    # === 工况条件（4列）===
    ("输入条件",          18),
    ("协议",              14),
    ("输出电压(V)",       14),
    ("输出电流(A)",       14),

    # === 实测值（3列）===
    ("最大值",            13),
    ("最小值",            13),
    ("有效值",            13),

    # === 纹波 & 能效（3列）===
    ("纹波要求",          13),
    ("纹波实测值(mV)",    16),
    ("效率(%)",           10),

    # === 结论（2列）===
    ("测试结论",          12),
    ("备注",              32),
]

def _get_cols(case_name_en: str = None):
    """
    获取指定用例的列定义。
    统一走 GLOBAL_COLS（所有用例共享同一套列），
    未注册用例也走 GLOBAL_COLS。
    """
    return GLOBAL_COLS

# CASE_COLS 保持 dict 接口兼容（_write_case_sheet 用 case_name 做 key 查询）
# = None 表示用 GLOBAL_COLS 作为所有用例的列模板
CASE_COLS = None


# ============================================================
# 样式
# ============================================================
def _mkstyle():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return dict(
        border=border,
        hdr_fill=PatternFill("solid", fgColor="1976D2"),
        hdr_font=Font(bold=True, color="FFFFFF", size=10),
        title_font=Font(bold=True, size=13, color="212121"),
        sub_font=Font(size=9, color="757575"),
        pass_fill=PatternFill("solid", fgColor="C8E6C9"),
        fail_fill=PatternFill("solid", fgColor="FFCDD2"),
        skip_fill=PatternFill("solid", fgColor="F5F5F5"),
        alt_fill=PatternFill("solid", fgColor="FAFAFA"),
    )


def _rfill(result: str, s: dict):
    return {"PASS": s["pass_fill"], "FAIL": s["fail_fill"], "SKIP": s["skip_fill"]}.get(result.upper())


def _fmt(v):
    if v == "" or v is None:
        return ""
    try:
        return f"{float(v):.3f}"
    except (ValueError, TypeError):
        return str(v)


# ============================================================
# 数据展开
# ============================================================

def _warn_missing_fields(name_en: str, sr: dict):
    """
    检查 sub_result 是否包含 GLOBAL_COLS 中定义的标准字段（可选）。
    由于不同用例使用不同字段子集，不再强制要求。
    此函数在新架构下已不再调用，保留以备将来审计用。
    """
    pass  # 新架构下不做强制检查，用例自行负责输出正确字段


def _flatten(r: dict):
    """
    将一个 result（一个测试用例）的 sub_results 展开为行列表。

    统一策略：sub_result 的字段名 = GLOBAL_COLS 列头（完全一致），
    不再进行任何映射转换。用例负责输出正确字段名。

    不在 GLOBAL_COLS 中的字段会被 _write_case_sheet 自动忽略。
    """
    name_en = r.get("name", "")
    name_cn = _cn(name_en)
    case_result = r.get("result", "")
    sub_results = r.get("sub_results", [])
    rows = []
    seq = 0
    if sub_results:
        for sr in sub_results:
            seq += 1
            skipped = sr.get("skipped", False)
            sub_pass = sr.get("overall_pass", None)
            if skipped:
                conclusion = "SKIP"
            elif sub_pass is True:
                conclusion = "PASS"
            elif sub_pass is False:
                conclusion = "FAIL"
            else:
                conclusion = case_result

            # 剥离 sweep_points（大数组，不写入报告）
            sr_clean = {k: v for k, v in sr.items() if k != "sweep_points"}

            # 统一字段：序号、用例名称（中文）和测试结论由 _flatten 统一注入
            row = {
                "序号": seq,
                "用例名称": name_cn,
                "测试结论": conclusion,
            }

            # 用例直接输出的字段（字段名即 GLOBAL_COLS 列头）
            row.update(sr_clean)

            rows.append(row)
    return rows


def generate_excel(results_path: str, output_dir: str = None,
                   dut_name: str = "") -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage

    with open(results_path, encoding="utf-8") as f:
        data = _json.load(f)

    summary     = data.get("summary", {})
    results     = data.get("results", [])
    export_time = data.get("export_time", "")

    # ---- 为每个 result 分配稳定 UUID（替代 id(result)，避免对象销毁后复用） ----
    for r in results:
        r.setdefault("_result_id", str(uuid.uuid4()))

    results_base = os.path.dirname(results_path)
    date_str = datetime.now().strftime("%Y%m%d_%H%M%S")

    def clean(s):
        return "".join(c for c in s if c not in r'\/:*?"<>|').strip()

    name_part = clean(dut_name) or "DUT"
    report_dir = output_dir or results_base
    os.makedirs(report_dir, exist_ok=True)
    file_name = f"{date_str}_{name_part}.xlsx"
    output_path = os.path.join(report_dir, file_name)

    s = _mkstyle()
    wb = Workbook()

    # -------------------------------------------------------
    # Sheet 1: 汇总
    # -------------------------------------------------------
    ws_sum = wb.active
    ws_sum.title = "测试结果汇总"
    ws_sum.column_dimensions["A"].width = 26
    ws_sum.column_dimensions["B"].width = 16
    ws_sum.column_dimensions["C"].width = 16
    ws_sum.column_dimensions["D"].width = 16
    ws_sum.column_dimensions["E"].width = 12

    ws_sum.merge_cells("A1:E1")
    t = ws_sum["A1"]
    t.value = "电源自动化测试平台 - 测试报告"
    t.font = s["title_font"]
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 28

    ws_sum.merge_cells("A2:E2")
    ws_sum["A2"].value = f"生成时间: {export_time}    产品: {name_part}"
    ws_sum["A2"].font = s["sub_font"]
    ws_sum["A2"].alignment = Alignment(horizontal="center")

    cat_stats = defaultdict(lambda: {"total": 0, "passed": 0, "failed": 0, "skipped": 0})
    for r in results:
        cn = _cn(r.get("name", ""))
        cat = _cat(cn)
        cat_stats[cat]["total"] += 1
        res = r.get("result", "")
        if res == "PASS":   cat_stats[cat]["passed"] += 1
        elif res == "FAIL": cat_stats[cat]["failed"] += 1
        elif res == "SKIP": cat_stats[cat]["skipped"] += 1

    row = 4
    ws_sum.merge_cells(f"A{row}:E{row}")
    ws_sum[f"A{row}"].value = "▌ 测试汇总"
    ws_sum[f"A{row}"].font = Font(bold=True, size=11, color="1976D2")
    row += 1

    for col_i, h in enumerate(["指标", "数值"], 1):
        c = ws_sum.cell(row, col_i, h)
        c.font = s["hdr_font"]; c.fill = s["hdr_fill"]
        c.border = s["border"]; c.alignment = Alignment(horizontal="center")
    ws_sum.merge_cells(f"B{row}:E{row}")
    row += 1

    for label, val in [
        ("总用例数", summary.get("total", 0)),
        ("通过",     summary.get("passed", 0)),
        ("失败",     summary.get("failed", 0)),
        ("跳过",     summary.get("skipped", 0)),
        ("通过率",   summary.get("pass_rate", "0%")),
        ("执行状态", summary.get("state", "")),
    ]:
        ws_sum.cell(row, 1, label).border = s["border"]
        ws_sum.cell(row, 1).font = Font(bold=True, size=9)
        ws_sum.merge_cells(f"B{row}:E{row}")
        vc = ws_sum.cell(row, 2, val)
        vc.border = s["border"]
        if label == "通过":   vc.fill = s["pass_fill"]
        elif label == "失败": vc.fill = s["fail_fill"]
        row += 1

    row += 1
    ws_sum.merge_cells(f"A{row}:E{row}")
    ws_sum[f"A{row}"].value = "▌ 分类统计"
    ws_sum[f"A{row}"].font = Font(bold=True, size=11, color="1976D2")
    row += 1

    for col_i, h in enumerate(["测试分类", "总数", "通过", "失败", "跳过"], 1):
        c = ws_sum.cell(row, col_i, h)
        c.font = s["hdr_font"]; c.fill = s["hdr_fill"]
        c.border = s["border"]; c.alignment = Alignment(horizontal="center")
    row += 1

    for cat in ["输入测试", "输出测试", "保护功能测试", "协议测试", "极限测试"]:
        if cat not in cat_stats:
            continue
        st = cat_stats[cat]
        t, p, f, sk = st["total"], st["passed"], st["failed"], st["skipped"]
        rate = f"{p/t*100:.1f}%" if t > 0 else "0%"
        for col_i, v in enumerate([cat, str(t), f"{p}({rate})", str(f), str(sk)], 1):
            c = ws_sum.cell(row, col_i, v)
            c.border = s["border"]
            c.alignment = Alignment(horizontal="center" if col_i > 1 else "left")
            if col_i == 3 and p > 0: c.fill = s["pass_fill"]
            elif col_i == 4 and f > 0: c.fill = s["fail_fill"]
        row += 1

    # -------------------------------------------------------
    # 波形 Sheet（先写，获取每用例起始行）
    # all_wf_entries: [(wf_path, disp_name, case_name), ...]
    # case_start_info: [(global_wf_idx, case_name, result_id), ...]
    # -------------------------------------------------------
    all_wf_entries = []
    case_start_info = []
    prev_rid = None

    # 波形自动查找：从 results/测试波形 目录根据用例名+输入条件+协议查找 PNG
    import glob as _glob
    def _wf_discovery(case_name_en, input_cond, proto_label, results_base, label_suffix=""):
        wf_dir = os.path.join(results_base, "测试波形")
        if not os.path.isdir(wf_dir):
            return ""
        # label_suffix 如 "_startup" 或 "_shutdown"，使 pattern 更精确地匹配对应波形
        pattern = f"{case_name_en}_{input_cond}_{proto_label}{label_suffix}.png"
        matches = _glob.glob(os.path.join(wf_dir, pattern))
        if matches:
            return max(matches, key=os.path.getmtime)
        return ""

    for r in results:
        flat_rows = _flatten(r)
        case_name_en = r.get("name", "")
        r_id = r.get("_result_id")
        case_nm = flat_rows[0].get("用例名称", "") if flat_rows else ""
        wf_start_idx_before_result = len(all_wf_entries)
        for row_data in flat_rows:
            # 只支持单波形列（测试波形）
            wf_keys = ["测试波形", "waveform"]
            for wf_key in wf_keys:
                wf = row_data.get(wf_key, "")
                # 尝试自动从测试波形目录查找
                if not wf or not os.path.isfile(wf):
                    wf = _wf_discovery(
                        case_name_en,
                        row_data.get("输入条件", ""),
                        row_data.get("协议", ""),
                        os.path.dirname(results_path),
                        "",
                    )
                if wf and os.path.isfile(wf):
                    disp = os.path.basename(wf)
                    all_wf_entries.append((wf, disp, case_nm))
                    break  # 只取第一个找到的波形
        # 只有本 result 至少有一个波形时才记录（跳过无波形的 result；避免空 flat_row 导致 this_case_count 错误）
        if len(all_wf_entries) > wf_start_idx_before_result:
            case_start_info.append((wf_start_idx_before_result, case_nm, r_id))
            prev_rid = r_id

    if all_wf_entries:
        ws_wf = wb.create_sheet("测试波形")
        case_rows = _write_waveform_sheet(ws_wf, all_wf_entries, case_start_info, s, XLImage, PILImage)
    else:
        case_rows = {}

    # ---- 建立 result_id -> wave_offset 映射 ----
    result_wf_offset = {}
    prev_rid = None
    for wf_idx, (wf_path, disp_name, case_nm) in enumerate(all_wf_entries):
        for info_idx, (ginfo_idx, info_case_nm, info_rid) in enumerate(case_start_info):
            if ginfo_idx == wf_idx and info_rid != prev_rid:
                result_wf_offset[info_rid] = wf_idx
                prev_rid = info_rid
                break

    # -------------------------------------------------------
    # 每个测试用例（result）建一个独立 Sheet
    # -------------------------------------------------------
    for r in results:
        flat_rows = _flatten(r)
        cat = _cat(_cn(r.get("name", "")))
        r_id = r.get("_result_id")

        raw_name = r.get("name", "")
        sheet_base = _cn(raw_name) if raw_name else cat
        sheet_name = "".join(c for c in sheet_base if c not in r'\/:*?"<>|').strip()
        if not sheet_name:
            sheet_name = cat

        existing = [ws.title for ws in wb.worksheets]
        i = 1
        while sheet_name in existing:
            sheet_name = f"{sheet_base}({i})"
            sheet_name = "".join(c for c in sheet_name if c not in r'\/:*?"<>|').strip()
            i += 1

        ws = wb.create_sheet(sheet_name)
        _write_case_sheet(ws, r, flat_rows, cat, s,
                          result_wf_offset.get(r_id, 0), case_rows,
                          results_base=os.path.dirname(results_path))

    wb.save(output_path)

    print("Excel report generated: " + output_path)
    return output_path


# ============================================================
# ============================================================
# 用例列定义查询
# 每个测试用例在类级别定义 COLS = [(列名, 宽度), ...]，顺序即报告列顺序
# ============================================================

def _get_case_cols(case_name: str):
    """
    根据用例名，从 CASE_REGISTRY 找到对应的类，读取 COLS 属性。
    返回：(cols_list, static_cols)
    - cols_list:  [(列名, 宽度), ...]  按用例定义的顺序
    - static_cols: 保持向后兼容返回空 set，实际不使用
    如果类未定义 COLS，fallback 到 GLOBAL_COLS（保持向后兼容）。

    实现：直接读取源码文件，正则提取 COLS 定义，绕过 import 链
    （避免 numpy 版本不兼容导致的包导入失败）。
    """
    try:
        entry = TestEngine.CASE_REGISTRY.get(case_name, {})
        if not entry:
            raise ValueError(f"{case_name} not in CASE_REGISTRY")
        mod_str = entry.get("module", "") if isinstance(entry, dict) else str(entry)
        if not mod_str:
            raise ValueError(f"{case_name} has no module key")
        parts = mod_str.split(".")
        if len(parts) != 3:
            raise ValueError(f"bad path: {module_path}")

        # ---- 方案：直接读源码文件，正则提取 COLS ----
        import os, re
        # 项目根目录（已知固定路径）
        _prj_root = r'D:\injoinic--job\自动化测试平台开发\自动化测试平台'
        py_file = os.path.join(_prj_root, *parts) + ".py"
        if os.path.isfile(py_file):
            with open(py_file, encoding="utf-8") as fh:
                src = fh.read()
            # 提取 COLS = [...] 块（括号平衡算法）
            m = re.search(r'COLS\s*=\s*\[', src)
            if m:
                start = m.start()
                depth = 0
                for i, ch in enumerate(src[start:], start):
                    if ch == '[':
                        depth += 1
                    elif ch == ']':
                        depth -= 1
                        if depth == 0:
                            cols_block = src[start:i+1]
                            break
                # 提取每个 (列名, 宽度) 元组
                tuples = re.findall(r'\("([^"]+)"\s*,\s*(\d+)\)', cols_block)
                if tuples:
                    cols = [(name, int(w)) for name, w in tuples]
                    return cols, set()
    except Exception:
        pass
    # Fallback：未知用例（未在 CASE_REGISTRY 中）或类未定义 COLS 属性时，
    # 使用 GLOBAL_COLS（保持向后兼容，通常是还未迁移到 COLS 体系的老用例）
    return GLOBAL_COLS, set()


def _write_case_sheet(ws, result: dict, sub_rows: list, cat: str, s: dict,
                       wf_offset: int, case_rows: dict = None, *, results_base: str = ""):
    """
    写入一个测试用例的独立 Sheet。

    列顺序和列宽：完全由用例的 COLS 属性定义，按用例定义顺序渲染。
    """
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    case_name = result.get("name", "")
    all_cols, _ = _get_case_cols(case_name)
    cols = list(all_cols)   # [(col_name, width), ...] 按用例定义顺序
    N = len(cols)
    col_keys = [cn for cn, _ in cols]

    # ---- ���前端列：号号 + 测试用例名称（不依起各 case COLS 定义） ----
    prefix_cols = [
        ("序号", 6),
        ("用例名称", 14),
    ]
    prefix_keys = [k for k, _ in prefix_cols]
    all_render_cols = prefix_cols + cols
    all_render_keys = prefix_keys + col_keys
    N_total = len(all_render_cols)

    # ---- 步骤2：列宽 ----
    for col_i, (col_name, w) in enumerate(all_render_cols, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = w

    # ---- 第1行：标题 ----
    case_title = sub_rows[0].get("用例名称", cat) if sub_rows else cat
    ws.row_dimensions[1].height = 24
    ws.merge_cells(f"A1:{get_column_letter(N_total)}1")
    t = ws["A1"]
    t.value = f"{case_title} - 测试明细"
    t.font = s["title_font"]
    t.alignment = Alignment(horizontal="center", vertical="center")

    # ---- 第2行：列头 ----
    ws.row_dimensions[2].height = 24
    for col_i, (col_name, _) in enumerate(all_render_cols, 1):
        c = ws.cell(2, col_i, col_name)
        c.font = s["hdr_font"]
        c.fill = s["hdr_fill"]
        c.border = s["border"]
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A3"

    wave_sheet_name = "测试波形"
    cur_row = 3
    running_wf_idx = 0

    # 列头名 -> JSON key 映射
    #
    # 规范（2026-04-23）：_make_result 返回的 key 必须与 COLS 列头完全一致，
    # 报告生成器直接用列头名查 JSON，不再需要 alias 转换。
    _key_alias = {}  # 直接用列头名查 JSON

    # ---- 自动查找波形路径（JSON 无波形时，从 results 目录的备份子目录查找）----
    def _find_waveform_png(case_name: str, input_cond: str, proto_label: str, label_suffix: str = "") -> str:
        """
        根据用例名+输入条件+协议+波形标签，从 results/测试波形 目录查找 PNG 文件。
        label_suffix 如 "_startup" / "_shutdown"，使 pattern 更精确匹配。
        返回找到的完整路径，不存在返回空字符串。
        """
        import glob as _glob
        wf_dir = os.path.join(results_base, "测试波形")
        if not os.path.isdir(wf_dir):
            return ""
        pattern = f"{case_name}_{input_cond}_{proto_label}{label_suffix}.png"
        matches = _glob.glob(os.path.join(wf_dir, pattern))
        if matches:
            return max(matches, key=os.path.getmtime)
        return ""

    for row_data in sub_rows:
        conclusion = row_data.get("测试结论", "")
        ws.row_dimensions[cur_row].height = 18

        for col_i, col_name in enumerate(all_render_keys, 1):
            if col_name == "序号":
                v = cur_row - 2
            elif col_name == "用例名称":
                v = row_data.get("用例名称", "")
            elif col_name == "测试结论":
                v = conclusion
            elif col_name == "测试波形":
                # 优先用 JSON 里已有的波形路径，其次从测试波形目录自动查找
                wf_path = row_data.get("测试波形", "")
                if not wf_path or not os.path.isfile(wf_path):
                    case_nm_en = result.get("name", "")
                    wf_path = _find_waveform_png(case_nm_en,
                        row_data.get("输入条件", ""),
                        row_data.get("协议", ""))
                v = wf_path  # 写入完整路径供后续超链接使用
            elif col_name in ("开机波形", "关机波形"):
                # 从 JSON key 直接读取（_make_result 已写入），不做 auto-discovery
                wf_path = row_data.get(col_name, "")
                v = wf_path
            else:
                # 列头名即 COLS 定义；通过映射找 JSON key，找不到则用列头名本身
                json_key = _key_alias.get(col_name, col_name)
                raw_v = row_data.get(json_key, "")
                if raw_v != "" and raw_v is not None:
                    try:
                        raw_v = f"{float(raw_v):.3f}"
                    except (ValueError, TypeError):
                        raw_v = str(raw_v)
                v = raw_v if raw_v != "" else "-"

            c = ws.cell(cur_row, col_i, v)
            c.border = s["border"]
            if col_i == 1:
                c.alignment = Alignment(horizontal="center")
            elif col_name == "测试结论":
                c.alignment = Alignment(horizontal="center", vertical="center")
                rf = _rfill(conclusion, s)
                if rf:
                    c.fill = rf
            elif col_name in ("测试波形",):
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")

        # ---- 超链接：每个 flat_row 只有 1 张波形（测试波形列）----
        # 必须用 all_render_keys（含 prefix_keys 的"序号"+"测试用例名称"），
        # 才能得到正确的列索引（1-based，与 ws.cell() 对齐）
        wf_col_idx = None
        for wf_col_name in ("测试波形",):
            if wf_col_name in all_render_keys:
                wf_col_idx = all_render_keys.index(wf_col_name) + 1
                wf_json_key = wf_col_name
                break

        if wf_col_idx and case_rows:
            actual_wf = row_data.get(wf_json_key, "") if wf_json_key else ""
            has_wf = bool(actual_wf and os.path.isfile(actual_wf))

            if has_wf:
                # 每个 flat_row 对应 all_wf_entries 中 1 个波形
                global_wf_idx = wf_offset + running_wf_idx
                case_start_wf_idx = 0
                for start_idx in sorted(case_rows.keys(), reverse=True):
                    if start_idx <= global_wf_idx:
                        case_start_wf_idx = start_idx
                        break
                case_start_row = case_rows.get(case_start_wf_idx, 2)
                pos_in_case = global_wf_idx - case_start_wf_idx
                name_row = case_start_row + 1 + (pos_in_case // 4) * 2
                col_in_case = pos_in_case % 4
                col_letter = get_column_letter(col_in_case + 1)
                wf_name_cell_addr = f"{col_letter}{name_row}"

                wave_link_cell = ws.cell(cur_row, wf_col_idx)
                from openpyxl.worksheet.hyperlink import Hyperlink
                wave_link_cell.hyperlink = Hyperlink(
                    ref=wave_link_cell.coordinate,
                    location=f"{wave_sheet_name}!{wf_name_cell_addr}",
                    display="波形"
                )
                wave_link_cell.value = "波形"
                wave_link_cell.font = Font(color="1976D2", underline="single", size=9)
                wave_link_cell.alignment = Alignment(horizontal="center", vertical="center")
                wave_link_cell.border = s["border"]

                running_wf_idx += 1

        if (cur_row - 2) % 2 == 0:
            for col_i, col_name in enumerate(col_keys, 1):
                if col_name == "测试结论":
                    continue
                ws.cell(cur_row, col_i).fill = s["alt_fill"]
        cur_row += 1

    # ============================================================
    # InputEfficiencyTest：合并 6级/7级平均能效列的 5 行负载点格子
    # ============================================================
    if result.get("name") == "InputEfficiencyTest" and sub_rows:
        _merge_efficiency_avg_cells(ws, sub_rows, col_keys, s)


def _merge_efficiency_avg_cells(ws, sub_rows, col_keys, s):
    """
    每个测试条件组（5 个负载点行）内，将 6级/7级平均能效列各自合并为 1 个格。

    6级平均：取 50% 负载行数据，写入该组第 1 格的合并格
    7级平均：取 100% 负载行数据，写入该组第 1 格的合并格
    """
    from openpyxl.styles import Alignment

    # 找 6级 和 7级 列索引（1-based）
    col_6l = None
    col_7l = None
    for ci, cn in enumerate(col_keys, 1):
        if cn == "6级平均能效(%)":
            col_6l = ci
        if cn == "7级平均能效(%)":
            col_7l = ci
    if col_6l is None and col_7l is None:
        return

    data_start_row = 3  # 数据从第3行开始
    ROWS_PER_COND = 5   # 每个测试条件 5 个负载点
    n = len(sub_rows)

    # 按测试条件分组（每 5 行为一组）
    num_conds = n // ROWS_PER_COND  # 有多少个测试条件

    for cond_idx in range(num_conds):
        group_start = cond_idx * ROWS_PER_COND  # 组内第一行在 sub_rows 的索引
        group_rows = sub_rows[group_start: group_start + ROWS_PER_COND]

        # 该组在 Excel 中的行号（从 data_start_row 起）
        excel_start = data_start_row + group_start
        excel_end   = excel_start + ROWS_PER_COND - 1

        # 从组内找 6级（50%）和 7级（100%）avg 值和结论
        val_6l = val_7l = None
        pass_6l = pass_7l = ""
        for sr in group_rows:
            lp = str(sr.get("负载点", ""))
            if lp == "50%" and col_6l:
                v = sr.get("6级平均能效(%)", "")
                if v and v != "-" and float(v) > 0:
                    val_6l = v
                pass_6l = sr.get("6级能效结论", "") or ""
            if lp == "100%" and col_7l:
                v = sr.get("7级平均能效(%)", "")
                if v and v != "-" and float(v) > 0:
                    val_7l = v
                pass_7l = sr.get("7级能效结论", "") or ""

        # 合并 6级平均能效列
        if col_6l and val_6l is not None:
            try:
                v_display = f"{float(val_6l):.2f}"
            except (ValueError, TypeError):
                v_display = str(val_6l)
            ws.merge_cells(
                start_row=excel_start, start_column=col_6l,
                end_row=excel_end, end_column=col_6l
            )
            c = ws.cell(excel_start, col_6l)
            c.value = v_display
            c.alignment = Alignment(horizontal="center", vertical="center")

        # 合并 7级平均能效列
        if col_7l and val_7l is not None:
            try:
                v_display = f"{float(val_7l):.2f}"
            except (ValueError, TypeError):
                v_display = str(val_7l)
            ws.merge_cells(
                start_row=excel_start, start_column=col_7l,
                end_row=excel_end, end_column=col_7l
            )
            c = ws.cell(excel_start, col_7l)
            c.value = v_display
            c.alignment = Alignment(horizontal="center", vertical="center")

        # 合并 6级/7级能效结论列（分别找对应列索引）
        col_c6l = col_keys.index("6级能效结论") + 1 if "6级能效结论" in col_keys else None
        col_c7l = col_keys.index("7级能效结论") + 1 if "7级能效结论" in col_keys else None

        if col_c6l and pass_6l:
            ws.merge_cells(
                start_row=excel_start, start_column=col_c6l,
                end_row=excel_end, end_column=col_c6l
            )
            c = ws.cell(excel_start, col_c6l)
            c.value = pass_6l
            c.alignment = Alignment(horizontal="center", vertical="center")

        if col_c7l and pass_7l:
            ws.merge_cells(
                start_row=excel_start, start_column=col_c7l,
                end_row=excel_end, end_column=col_c7l
            )
            c = ws.cell(excel_start, col_c7l)
            c.value = pass_7l
            c.alignment = Alignment(horizontal="center", vertical="center")


# ============================================================
# 波形 Sheet（新布局：4列 × 每用例3行 × 4波形/行 × 用例间空4行）
# ============================================================

def _write_waveform_sheet(ws, all_wf_entries: list, case_start_info: list,
                           s: dict, XLImage, PILImage):
    """
    all_wf_entries:   [(wf_path, disp_name, case_name), ...] 全局扁平波形（按分类顺序）
    case_start_info: [(global_wf_idx, case_name, result_id), ...] 每个用例首波形的索引、名称和result_id
    布局：每个 case 占 3 行（case名行 / 波形名行 / 波形行），用例间空 4 行。
    返回: {global_wf_idx: case_start_row}  用于超链接目标计算
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font

    if not all_wf_entries:
        return {}

    # ---- 固定4列，列宽50 ----
    for col_i in range(1, 5):
        ws.column_dimensions[get_column_letter(col_i)].width = 50

    # ---- 第1行：标题 ----
    ws.row_dimensions[1].height = 24
    ws.merge_cells('A1:D1')
    t = ws['A1']
    t.value = '测试波形'
    t.font = s['title_font']
    t.alignment = Alignment(horizontal='center', vertical='center')

    SCALE = 0.5
    case_rows = {}    # {global_wf_idx_of_case_first_waveform: case_start_row}
    cur_row = 2      # 波形从第2行开始

    # 按 case 分组处理 all_wf_entries
    case_info_iter = iter(case_start_info)
    next_case_start = next(case_info_iter, None)   # (global_wf_idx, case_name, result_id)
    wf_idx = 0
    n_entries = len(all_wf_entries)

    while wf_idx < n_entries:
        if next_case_start is None:
            break
        if wf_idx > 0:
            cur_row += 4   # 用例间空4行

        case_start_wf_idx = wf_idx
        case_start_row = cur_row
        case_rows[case_start_wf_idx] = case_start_row
        case_name = next_case_start[1]

        # ---- 确定当前 case 的波形数量 ----
        # case_start_info 每个条目对应 all_wf_entries 中的一个 waveform；
        # 同一个 flat_row 的 startup/shutdown 共享同一个 case 条目（第一个 waveform 的位置）。
        # 找下一个 case 条目的 ginfo_idx 即可确定 this_case_count。
        current_csi_idx = 0
        for i_csi, (csi_gidx, _, _) in enumerate(case_start_info):
            if csi_gidx == wf_idx:
                current_csi_idx = i_csi
                break
        next_csi_gidx = (case_start_info[current_csi_idx + 1][0]
                         if current_csi_idx + 1 < len(case_start_info)
                         else n_entries)
        this_case_count = next_csi_gidx - wf_idx

        # ---- 写 case name 行 ----
        case_name_row = cur_row
        ws.row_dimensions[case_name_row].height = 20
        c = ws.cell(case_name_row, 1, case_name)
        c.font = Font(size=10, bold=True, color='1976D2')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = s['border']
        ws.merge_cells(f'A{case_name_row}:D{case_name_row}')

        # ---- 波形名称行 + 波形行 ----
        case_waveforms = all_wf_entries[wf_idx:wf_idx + this_case_count]
        # 每 4 张波形换一行，避免重叠
        COLS_PER_ROW = 4
        wf_group_idx = 0
        while wf_group_idx < this_case_count:
            group_start = wf_group_idx
            group_end   = min(wf_group_idx + COLS_PER_ROW, this_case_count)
            group_size  = group_end - wf_group_idx

            wf_name_row = case_name_row + 1 + (wf_group_idx // COLS_PER_ROW) * 2
            wf_img_row  = wf_name_row + 1
            ws.row_dimensions[wf_name_row].height = 20
            ws.row_dimensions[wf_img_row].height  = 250

            for local_idx in range(group_start, group_end):
                img_path, disp_name, _ = case_waveforms[local_idx]
                col = (local_idx % COLS_PER_ROW) + 1
                col_letter = get_column_letter(col)

                name_cell = ws.cell(wf_name_row, col, disp_name)
                name_cell.font = Font(size=9, bold=False, color='1976D2')
                name_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                name_cell.border = s['border']

                try:
                    with PILImage.open(img_path) as im:
                        orig_w, orig_h = im.size
                except:
                    orig_w, orig_h = 800, 600
                img = XLImage(img_path)
                img.width  = int(orig_w * SCALE)
                img.height = int(orig_h * SCALE)
                img.anchor = f'{col_letter}{wf_img_row}'
                ws.add_image(img)

            wf_group_idx += COLS_PER_ROW

        wf_idx += this_case_count
        cur_row = case_name_row + 1 + ((this_case_count - 1) // COLS_PER_ROW + 1) * 2
        next_case_start = next(case_info_iter, None)

    return case_rows
# ============================================================
# rels 绝对路径修复（openpyxl → 相对路径）
# ============================================================
# 简化版 rels 修复：仅处理内部超链接的 rels 条目创建
# ============================================================

def _fix_rels(xlsx_path: str):
    """
    openpyxl 3.x 的 Hyperlink 对象会生成 r:id，但不创建对应的 rels 条目。
    本函数遍历所有 worksheet，为内部超链接创建缺失的 rels 条目。
    """
    import shutil, zipfile, re

    # 读取 ZIP 内容
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        names = z.namelist()
        files = {name: z.read(name) for name in names}

    new_files = {}  # filename -> content

    # 遍历所有 worksheet，找到有超链接的
    for name in names:
        if not name.startswith('xl/worksheets/sheet') or not name.endswith('.xml'):
            continue
        if b'<hyperlink' not in files[name]:
            continue

        # 计算对应的 rels 文件路径
        # xl/worksheets/sheet3.xml -> xl/worksheets/_rels/sheet3.xml.rels
        parts = name.rsplit('/', 1)
        rels_name = parts[0] + '/_rels/' + parts[1] + '.rels'

        # 提取所有超链接的 location（bytes模式）
        raw = files[name]
        hyperlink_locs = re.findall(b'<hyperlink[^>]*location="([^"]+)"[^>]*/>', raw)
        if not hyperlink_locs:
            continue

        # 收集已有的非超链接 rels 条目
        existing_rels = files.get(rels_name, b'')
        existing_entries = []
        if existing_rels:
            pos = 0
            while True:
                elem_start = existing_rels.find(b'<Relationship', pos)
                if elem_start < 0:
                    break
                elem_end = existing_rels.find(b'/>', elem_start)
                if elem_end < 0:
                    break
                elem_end += 2
                elem = existing_rels[elem_start:elem_end]
                # 保留非超链接的 relationship
                if b'hyperlink' not in elem:
                    existing_entries.append(elem)
                pos = elem_end

        # 为每个 location 创建 hyper link rels 条目
        rid_counter = 1
        hyperlink_entries = []
        for loc_bytes in hyperlink_locs:
            rid = f'rId{rid_counter}'
            rid_counter += 1
            loc_str = loc_bytes.decode('utf-8', errors='replace')
            entry = ('<Relationship Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink" '
                     f'Target="{loc_str}" TargetMode="Internal" Id="{rid}"/>').encode('utf-8')
            hyperlink_entries.append(entry)

        # 构建新的 rels 内容
        all_entries = existing_entries + hyperlink_entries
        new_rels = b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">' + b''.join(all_entries) + b'</Relationships>'
        new_files[rels_name] = new_rels

    # Step 2: 修复图片路径（绝对路径 -> 相对路径）
    for name in names:
        if not name.endswith('.rels'):
            continue
        if name in new_files:
            continue  # 已在 Step 1 处理

        raw = files[name]
        changed = False

        # 修复绝对路径 /xl/media/ -> ../media/
        if b'Target="/xl/media/' in raw:
            raw = raw.replace(b'Target="/xl/media/', b'Target="../media/')
            changed = True

        # 移除不需要的 TargetMode="External"
        if b'TargetMode="External"' in raw:
            raw = raw.replace(b' TargetMode="External"', b'')
            changed = True

        if changed:
            new_files[name] = raw

    # 写入修改后的文件
    if new_files:
        tmp = xlsx_path + '.tmp'
        with zipfile.ZipFile(xlsx_path, 'r') as zin:
            original_names = set(zin.namelist())
            with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
                # 复制原有文件
                for item in zin.infolist():
                    if item.filename in new_files:
                        zout.writestr(item, new_files[item.filename])
                    else:
                        zout.writestr(item, zin.read(item.filename))
                # 写入新文件（如 sheet3.xml.rels）
                for fname, content in new_files.items():
                    if fname not in original_names:
                        zout.writestr(fname, content)
        shutil.move(tmp, xlsx_path)


# ============================================================
# rels 绝对路径修复（openpyxl → 相对路径）
# ============================================================
import re





def auto_generate(results_dir: str = None, dut_name: str = ""):
    if results_dir is None:
        results_dir = os.path.join(os.path.dirname(__file__), "results")
    json_files = []
    if os.path.isdir(results_dir):
        for f in os.listdir(results_dir):
            # 扫描 result 子目录下所有 JSON（测试结果或配置备份），取最新
            if f.endswith(".json"):
                json_files.append(os.path.join(results_dir, f))
    if not json_files:
        print("[ReportGenerator] results 目录没有找到测试结果文件")
        return None
    latest = max(json_files, key=os.path.getmtime)
    print(f"[ReportGenerator] 生成报告: {latest}")
    return generate_excel(latest, results_dir, dut_name)


if __name__ == "__main__":
    auto_generate()